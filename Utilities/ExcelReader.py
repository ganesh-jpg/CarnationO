import openpyxl

#Function to read Total Number of Rows
def get_rows(path, sheetName):
    worksheet = openpyxl.load_workbook(path)
    sheet = worksheet[sheetName]
    return sheet.max_row

#Function to read Total Number of Columns
def get_columns(path, sheetName):
    worksheet = openpyxl.load_workbook(path)
    sheet = worksheet[sheetName]
    return sheet.max_column

#Function to read Cell Data
def get_Cell_Data(path, sheetName, rowNo, columnNo):
    worksheet = openpyxl.load_workbook(path)
    sheet = worksheet[sheetName]
    return sheet.cell(row=rowNo, column=columnNo).value

#Function to set Cell Data
def set_Cell_Data(path, sheetName, rowNo, columnNo, Data):
    worksheet = openpyxl.load_workbook(path)
    sheet = worksheet[sheetName]
    sheet.cell(row=rowNo, column=columnNo).value = Data
    worksheet.save(path)

'''
#Function call to Test
path = "..\\Excel\\Test.xlsx"
sheetName = "Register"
print(get_rows(path,sheetName))
print(get_columns(path,sheetName))
rowNo = 2
columnNo = 1
print(get_Cell_Data(path,sheetName,rowNo,columnNo))
rowNo = 3
columnNo = 1
Data = "Monika"
set_Cell_Data(path,sheetName,rowNo,columnNo,Data)
'''